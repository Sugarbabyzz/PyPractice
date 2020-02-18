import pandas as pd
import os
from openpyxl import load_workbook

#  1、建立单位名称和序号的对应字典
df = pd.read_excel('单位名称对应序号.xlsx', index_col=None)
num_map_dic = {}
for i, r in df.iterrows():
    num_map_dic[r['单位名称']] = r['序号']
print(num_map_dic)


#  2、在表头写上序号
no_number_entity = []  # 无对应序号的单位
error_entity = []  # 出现错误的公司
count = 0

rootpath = 'data'  # 程序开始的文件夹
for street_dir in os.listdir(rootpath):
    street_path = rootpath + '/' + street_dir
    for com_dir in os.listdir(street_path):
        com_path = street_path + '/' + com_dir
        for filename in os.listdir(com_path):
            try:
                if '总表' in filename:
                    continue
                filepath = com_path + '/' + filename
                entity_name = filename[:-5]
                entity_number = num_map_dic.get(entity_name.replace('(', '（').replace(')', '）'))
                if entity_number is None:
                    no_number_entity.append(filepath)
                    continue
                print(filepath + '    ' + entity_name + ': ' + entity_number)
                wb = load_workbook(filepath)
                ws = wb.active
                ws['A52'] = entity_number
                wb.save(filepath)
                count += 1
            except:
                print('---------------------' + filepath + '  出现错误！ ---------------------')
                error_entity.append(filepath)


with open('processing_log.txt', 'a') as f:
    f.write('------------------------------------------------------------------------------------')
    f.write('统计如下：\n')
    f.write('已处理文件：')
    f.write(count)
    f.write('\n无对应序号的文件：')
    f.write(no_number_entity)
    f.write('\n出现错误的文件：')
    f.write(error_entity)





from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


'''
1、读取Excel
'''
# 默认可读写，若有需要可以指定write_only和read_only为True
wb = load_workbook('example.xlsx')
# 获得所有sheet的名称
print(wb.sheetnames)
# 根据sheet名字获得sheet
a_sheet = wb['Sheet1']
# 获得sheet名
print(a_sheet.title)

# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
sheet = wb.active
# 获得单元格内的内容
print(sheet['B4'].value)
# 获得最大行和列
print(sheet.max_row)
print(sheet.max_column)

# 获取行和列
# A1、B1、C1这样的顺序
for row in sheet.rows:
    for cell in row:
        print(cell.value)

# A1、A2、A3这样的顺序
for col in sheet.columns:
    for cell in col:
        print(cell.value)

# 获取指定的行列
for row in list(sheet.rows)[2]:  # sheet.rows是生成器类型，不能使用索引，转换成list之后再使用索引
    print(row.value)

# 获得任意区间的单元格
# A1到B3的矩形
for i in range(1, 4):
    for j in range(1, 3):
        print(sheet.cell(row=i, column=j))

for row_cell in sheet['A1':'B3']:  # 返回每行的元组
    for cell in row_cell:
        print(cell)

# 根据字母获得列号，根据列号返回字母
# 需要导入， 这两个函数存在于openpyxl.utils

# 根据列的数字返回字母
print(get_column_letter(2))  # B
# 根据字母返回列的数字
print(column_index_from_string('D'))  # 4

'''
2、写Excel
'''
from openpyxl import Workbook

# 工作表相关

wb = Workbook()
print(wb.sheetnames)  # 默认有一个Sheet
# 新建工作表，指定索引
wb.create_sheet('Data', index=1)  # 第一是index=0
# 删除工作表
# wb.remove('Data')
del wb['Data']
# 激活工作表
sheet = wb.active


# 单元格写入

# 直接给单元格赋值就行
sheet['A1'] = 'good'
print(sheet['A1'].value)
# B9处写入平均值
sheet['B9'] = '=AVERAGE(B2:B8)'

# append函数，只能按行写入
# 添加一行
row = [1, 2, 3, 4, 5]
sheet.append(row)
# 添加多行
rows = [
 ['Number', 'data1', 'data2'],
 [2, 40, 30],
 [3, 40, 25],
 [4, 50, 30],
 [5, 30, 10],
 [6, 25, 5],
 [7, 50, 10],
]
for row in rows:
    sheet.append(row)
# 按列写入
# print(list(zip(*rows)))
# sheet.append(list(zip(*rows)))


'''
3、保存
'''
# 设置单元格风格
from openpyxl.styles import Font, colors, Alignment

# 字体
bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)
sheet['A1'].font = bold_itatic_24_font

# 对齐方式
# 设置B1中的数据垂直居中和水平居中
sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

# 设置行高和列高
# 第2行行高
sheet.row_dimensions[2].height = 40
# C列列宽
sheet.column_dimensions['C'].width = 30

# 合并单元格
# 合并单元格， 往左上角写入数据即可
# 如果这些要合并的单元格都有数据，只会保留左上角的数据，其他则丢弃。
sheet.merge_cells('B1:G1')  # 合并一行中的几个单元格
sheet.merge_cells('A1:C3')  # 合并一个矩形区域中的单元格

# 拆分单元格
sheet.unmerge_cells('A1:C3')

# 保存
wb.save('pyxl_test.xlsx')













